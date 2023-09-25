from connector import *
from openpyxl import Workbook, load_workbook
from Operations import get_accounts_name
import json


def create_labels(headers, testId, aid, labels):

    labels_url = "https://api.thousandeyes.com/v6/groups/tests/new.json?aid=%s" %aid

    for label in labels:

        payload = json.dumps({"name": label,
                   "tests": [{
                            "testId": testId
                        }
                    ]
                })

        label_create = post_data(headers, endp_url=labels_url, payload=payload)

        if label_create:
            logging.info("Label created for " + str(testId) + " at " + str(aid) + " :" + label )
        else:
            logging.info("Label cannot be created for " + str(testId) + " at " + str(aid) + " :" + label )


        



def deploy_template(file_path, headers):

    wb = load_workbook(file_path)
    wb.active = wb["Deploy"] #para movernos a una pagina en específico
    ws = wb.active

    account2aid = get_accounts_name(headers)

    for row in ws.iter_rows(min_row=10, max_col=8, max_row=ws.max_row, values_only=True):
        

        if row[7] is None:
            continue

        accounts: list = row[6].split(", ")
        labels: list = row[7].split(", ")

        for account in accounts:

            aid = account2aid.get(account)
        
            if create_labels(headers, testId=row[4], aid=aid, labels=labels): #segun yo no necesitamos un if porque tenemos logs

                pass

